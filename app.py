from flask import Flask, render_template, request, send_file, session
from werkzeug.utils import secure_filename
import os
import webbrowser
from openpyxl import load_workbook
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Mm  # Para ajustar el tamaño de la imagen
import json
from cx_Freeze import setup, Executable
import sys


app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['GENERATED_FOLDER'] = 'generated'  # Carpeta para los archivos generados
app.config['SECRET_KEY'] = '39QDt7fVWUuPqLsPDAF3XkuDQEKiZkxN9z'
# Página principal para subir archivos
@app.route('/')
def index():
    return render_template('index.html')

# Función para formatear valores según el formato de la celda en Excel
def formatear_valor(celda):
    if isinstance(celda.value, (int, float)):
        # Si el valor es mayor o igual a 1, se formatea como dólares
        if celda.value >= 1:
            return "${:,.2f}".format(celda.value)  # Formato de dólares para valores positivos
        # Si el valor es negativo
        elif celda.value < 0:
            return "-${:,.2f}".format(abs(celda.value))  # Formato de dólares y paréntesis para valores negativos
        # Si el valor es un porcentaje
        elif celda.value >= 0:
            return "{:.2%}".format(celda.value)  # Formato de porcentaje para valores menores que 1
    else :
        return celda.value  # No aplicar formato si no es numérico

# Función para guardar el contexto en un archivo JSON
def guardar_contexto(contexto):
    with open('uploads/context.json', 'w') as f:
        json.dump(contexto, f)

# Función para cargar el contexto desde un archivo JSON
def cargar_contexto():
    if os.path.exists('uploads/context.json'):
        with open('uploads/context.json', 'r') as f:
            return json.load(f)
    return {}

# Subir archivos Excel (primera etapa)
@app.route('/upload', methods=['POST'])
def upload_files():
    if 'files' not in request.files:
        return 'No se han subido archivos.'

    files = request.files.getlist('files')
    filenames = []

    # Guardar archivos Excel subidos
    for file in files:
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        filenames.append(filename)

    # Mostrar opciones adicionales después de subir los archivos
    return render_template('index.html', filenames=filenames, show_additional_fields=True)

# Procesar texto, imágenes y generar archivo .docx (segunda etapa)
@app.route('/generate', methods=['POST'])
def generate_doc():
    # Cargar la plantilla .docx
    doc = DocxTemplate('static/1.docx')
    context = cargar_contexto()

    # Procesar archivos Excel cargados previamente
    uploaded_files = os.listdir(app.config['UPLOAD_FOLDER'])
    for idx, filename in enumerate(uploaded_files):
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        if filename.endswith('.xlsx'):
            # Usar openpyxl para leer el archivo y obtener los formatos
            wb = load_workbook(file_path, data_only=True)

            # Seleccionar la hoja correspondiente
            if 'ejemplo1' in filename:
                ws = wb.worksheets[1]  # Hoja 2
                evidencia_1 = formatear_valor(ws['D9'])  # Celda D9 con formato
                evidencia_2 = formatear_valor(ws['D10'])  # Celda D10 con formato
                context[f'ev1'] = evidencia_1
                context[f'ev2'] = evidencia_2

            elif 'ejemplo2' in filename:
                ws = wb.worksheets[2]  # Hoja 4
                evidencia_3 = formatear_valor(ws['B12'])  # Celda C11 con formato
                ws = wb.worksheets[3]  # Hoja 4
                evidencia_4 = formatear_valor(ws['C13'])  # Celda C12 con formato
                ws = wb.worksheets[5]  # Hoja 4
                evidencia_5 = formatear_valor(ws['C13'])
                evidencia_6 = formatear_valor(ws['K7'])
                evidencia_7 = formatear_valor(ws['K6'])
                evidencia_8 = formatear_valor(ws['F9'])

                context[f'ev3'] = evidencia_3
                context[f'ev4'] = evidencia_4
                context[f'ev5'] = evidencia_5
                context[f'ev6'] = evidencia_6
                context[f'ev7'] = evidencia_7
                context[f'ev8'] = evidencia_8

            # Si no es un archivo reconocido, usa la celda A1 por defecto
            else:
                evidencia_default = formatear_valor(ws['A1'])  # Celda A1 con formato
                context[f'evidencia_default_{idx}'] = evidencia_default

            # Eliminar el archivo Excel subido (opcional)
            os.remove(file_path)

    # Verificar el contenido del contexto para asegurarse de que todo esté correcto
    print("Contexto final:", context)

    # Rellenar la plantilla con los datos
    doc.render(context)

    # Guardar el nuevo archivo .docx lleno
    output_path = os.path.join(app.config['GENERATED_FOLDER'], 'informe_lleno.docx')
    
    # Eliminar archivo si ya existe
    if os.path.exists(output_path):
        os.remove(output_path)
    
    doc.save(output_path)

    # Enviar el archivo generado al cliente para su descarga
    return send_file(output_path, as_attachment=True)

def cargar_contexto():
    # Si quieres reiniciar siempre los valores al cargar, puedes devolver un diccionario vacío
    return {}

if __name__ == '__main__':
    webbrowser.open("http://127.0.0.1:5000") 
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])
    if not os.path.exists(app.config['GENERATED_FOLDER']):
        os.makedirs(app.config['GENERATED_FOLDER'])
    app.run(debug=True, use_reloader=False, threaded=False)
